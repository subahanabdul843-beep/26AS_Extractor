import streamlit as st
import pdfplumber
import pandas as pd
import re
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

st.set_page_config(page_title="26AS Extractor", layout="wide")

st.title("26AS PDF → Excel Extractor")

uploaded_file = st.file_uploader(
    "Upload Form 26AS PDF",
    type=["pdf"]
)

# ==============================================================
# HELPER FUNCTIONS
# ==============================================================

HEADER_FILL = PatternFill()
HEADER_FONT = Font(bold=True)
TOTAL_FONT  = Font(bold=True)
NUM_FORMAT  = '#,##0.00'

_table_counter = [0]          # mutable so nested helpers can bump it


def _next_table_name(prefix: str) -> str:
    """Return a unique Excel Table display-name."""
    _table_counter[0] += 1
    safe = re.sub(r'[^A-Za-z0-9]', '_', prefix)
    return f"{safe}_{_table_counter[0]}"


def apply_sheet_formatting(
    ws,
    numeric_cols: list = None,
    total_col_indices: list = None,
    table_name: str = None,
):
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=False)

    max_data_row = ws.max_row

    if total_col_indices:
        total_row = max_data_row + 1
        ws.cell(total_row, 1).value = "TOTAL"
        for col_idx, col_letter in total_col_indices:
            ws.cell(total_row, col_idx).value = (
                f"=SUM({col_letter}2:{col_letter}{max_data_row})"
            )
        for c in ws[total_row]:
            c.font = TOTAL_FONT
        max_data_row = total_row

    if numeric_cols:
        for col_idx in numeric_cols:
            for row in ws.iter_rows(min_row=2, max_row=max_data_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = NUM_FORMAT

    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 3
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 60)

    if table_name and ws.max_row > 1:
        max_col_letter = get_column_letter(ws.max_column)
        table_ref = f"A1:{max_col_letter}{ws.max_row}"
        tbl = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name=None,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        tbl.tableStyleInfo = style
        ws.add_table(tbl)


def write_df_sheet(writer, df: pd.DataFrame, sheet_name: str,
                    numeric_col_names: list = None,
                    add_totals: bool = False,
                    table_prefix: str = None,
                    placeholder_message: str = None):
    """Write a DataFrame (or a 'No Transactions Present' placeholder) to a sheet."""
    if df is None or df.empty:
        if placeholder_message:
            empty_df = pd.DataFrame([[placeholder_message]], columns=["Status"])
            empty_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            apply_sheet_formatting(ws)
        return

    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    col_map = {name: idx + 1 for idx, name in enumerate(df.columns)}

    num_indices = []
    if numeric_col_names:
        num_indices = [col_map[n] for n in numeric_col_names if n in col_map]

    total_pairs = []
    if add_totals and num_indices:
        total_pairs = [(i, get_column_letter(i)) for i in num_indices]

    tname = _next_table_name(table_prefix) if table_prefix else None

    apply_sheet_formatting(
        ws,
        numeric_cols=num_indices if num_indices else None,
        total_col_indices=total_pairs if total_pairs else None,
        table_name=tname,
    )


def safe_float(val) -> float:
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, AttributeError, TypeError):
        return 0.0


def safe_sheet_name(name: str) -> str:
    """Excel sheet names: <=31 chars, no : \\ / ? * [ ]"""
    safe = re.sub(r'[:\\/?*\[\]]', '', name)
    return safe[:31]


# ==============================================================
# DYNAMIC PART DETECTION ENGINE
# ==============================================================

PART_HEADING_RE = re.compile(r'^PART[\s-]*([IVXLCDM0-9]+)\s*[-\u2013]?\s*(.*)$')
NO_TRX_RE = re.compile(r'No\s+Transactions\s+Present', re.I)
END_MARKERS = ["Contact Information", "Legends used in Annual Tax Statement",
               "Notes for Annual Tax Statement", "*Notes:"]


def strip_noise(text: str, assessee_name: str = "") -> str:
    """Remove repeated page-header / page-footer artifact lines."""
    name_token = assessee_name.split()[-1].strip() if assessee_name else None
    lines = text.split("\n")
    out = []
    for line in lines:
        s = line.strip()
        if not s:
            continue
        if re.match(r'^Assessee PAN:\s*\S+\s+Assessee Name:.*Assessment Year:.*$', s):
            continue
        if re.match(r'^Data updated till .*$', s):
            continue
        if name_token and s == name_token:
            continue
        out.append(line)
    return "\n".join(out)


def find_parts(text: str):
    """Dynamically detect every 'PART-<token> - <Description>' heading and
    return the text block belonging to each Part, up to the next PART-
    heading or one of the known end-of-table markers. Not hardcoded to any
    specific set of Parts, so future Form 26AS layouts are supported."""
    lines = text.split("\n")
    headings = []

    for i, line in enumerate(lines):
        m = PART_HEADING_RE.match(line.strip())
        if m:
            roman = m.group(1)
            desc = m.group(2).strip(" -\u2013")
            if i + 1 < len(lines):
                nxt = lines[i + 1].strip()
                if nxt and not nxt.startswith("Sr.") and not PART_HEADING_RE.match(nxt) \
                        and not nxt.startswith("(All amount"):
                    desc = (desc + " " + nxt).strip()
            headings.append({"roman": roman, "description": desc, "line_idx": i})

    parts = []
    for idx, h in enumerate(headings):
        start_line = h["line_idx"]
        if idx + 1 < len(headings):
            end_line = headings[idx + 1]["line_idx"]
        else:
            end_line = len(lines)
            for j in range(start_line + 1, len(lines)):
                if any(lines[j].strip().startswith(mk) for mk in END_MARKERS):
                    end_line = j
                    break
        block = "\n".join(lines[start_line:end_line])
        parts.append({
            "roman": h["roman"],
            "label": f"Part-{h['roman']}",
            "description": h["description"],
            "text": block,
        })
    return parts


def has_no_transactions(part_text: str) -> bool:
    return bool(NO_TRX_RE.search(part_text))


def parse_contact_info(full_clean_text: str) -> dict:
    start = full_clean_text.find("Contact Information")
    if start == -1:
        return {}
    end = full_clean_text.find("Legends used in Annual Tax Statement", start)
    if end == -1:
        end = len(full_clean_text)
    block = full_clean_text[start:end]
    lines = [l.strip() for l in block.split("\n") if l.strip()]

    roman_re = re.compile(r'^([IVXLCDM]+)\s+(.*)$')
    contacts = {}
    current_roman = None
    for line in lines[2:]:
        m = roman_re.match(line)
        if m:
            current_roman = m.group(1)
            contacts[current_roman] = m.group(2).strip()
        elif current_roman:
            contacts[current_roman] = (contacts[current_roman] + " " + line).strip()

    return {k: v.rstrip("/").strip() for k, v in contacts.items()}


# ------------------------------------------------------------------
# Reusable field-type regex fragments, combined dynamically depending
# on which fields a given Part's table actually contains.
# ------------------------------------------------------------------
F_SNO     = r'(\d+)'
F_TAN     = r'([A-Z]{4}\d{5}[A-Z])'
F_DATE    = r'(\d{2}-[A-Za-z]{3}-\d{4})'
F_STATUS  = r'([A-Z])'
F_SECTION = r'([0-9]{1,4}[A-Z0-9]{0,5}(?:\([A-Za-z0-9]+\))*)'
F_AMOUNT  = r'(-?[\d,]+\.\d{2}|-?[\d,]+)'
F_REMARK  = r'(\S+)'
F_NAME    = r'([A-Z0-9 &.,()\'/-]+?)'

SUMMARY_HEADER_RE = re.compile(
    r'Name of (Deductor|Collector)\b.*TAN of (Deductor|Collector)', re.I
)
TRX_HEADER_RE = re.compile(r'Section\s*1?\s+Transaction Date', re.I)

summary_row_pattern = re.compile(
    rf'^{F_SNO}\s+{F_NAME}\s+{F_TAN}\s+{F_AMOUNT}\s+{F_AMOUNT}\s+{F_AMOUNT}\s*$'
)
trx_row_pattern = re.compile(
    rf'^{F_SNO}\s+{F_SECTION}\s+{F_DATE}\s+{F_STATUS}\s+{F_DATE}\s+{F_REMARK}\s+'
    rf'{F_AMOUNT}\s+{F_AMOUNT}\s+{F_AMOUNT}\s*$'
)


def is_deductor_collector_shape(part_text: str) -> bool:
    """Detects the recurring 'Name/TAN summary block followed by Section/
    Transaction Date detail rows' shape used by TDS- and TCS-style Parts
    (covers Part I, Part II, Part VI today, and any future Part sharing
    the same layout)."""
    return bool(SUMMARY_HEADER_RE.search(part_text)) and bool(TRX_HEADER_RE.search(part_text))


def parse_deductor_collector_shape(part_text: str):
    """Returns (summary_rows, transaction_rows) for the deductor/collector +
    transaction shape. Negative reversal rows, bracketed section codes
    (194I(a), 206C(1H) etc.), free-form Remarks tokens (G, A, B, F, *, ...),
    and multiple Sections per TAN are all supported."""
    summary_rows = []
    transaction_rows = []

    current_name = ""
    current_tan = ""
    section_lookup = {}

    role = "Deductor" if "deductor" in part_text[:400].lower() else "Collector"

    for line in part_text.split("\n"):
        line = line.strip()
        if not line:
            continue

        sm = summary_row_pattern.match(line)
        if sm:
            current_name = sm.group(2).strip()
            current_tan  = sm.group(3)
            summary_rows.append({
                "S.No": sm.group(1),
                f"Name of {role}": current_name,
                f"TAN of {role}": current_tan,
                "Section": "",   # filled in after full pass
                f"Total Amount Paid/{'Debited' if role == 'Collector' else 'Credited'}": safe_float(sm.group(4)),
                "Total Tax Deducted/Collected": safe_float(sm.group(5)),
                "Total Deposited": safe_float(sm.group(6)),
            })
            continue

        tm = trx_row_pattern.match(line)
        if tm:
            sec = tm.group(2)
            transaction_rows.append({
                f"Name of {role}": current_name,
                f"TAN of {role}": current_tan,
                "S.No": tm.group(1),
                "Section": sec,
                "Transaction Date": tm.group(3),
                "Status of Booking": tm.group(4),
                "Date of Booking": tm.group(5),
                "Remarks": tm.group(6),
                f"Amount Paid/{'Debited' if role == 'Collector' else 'Credited'}": safe_float(tm.group(7)),
                f"Tax {'Collected' if role == 'Collector' else 'Deducted'}": safe_float(tm.group(8)),
                f"{'TCS' if role == 'Collector' else 'TDS'} Deposited": safe_float(tm.group(9)),
            })
            if current_tan:
                section_lookup.setdefault(current_tan, [])
                if sec not in section_lookup[current_tan]:
                    section_lookup[current_tan].append(sec)

    section_str = {tan: ",".join(secs) for tan, secs in section_lookup.items()}
    tan_key = f"TAN of {role}"
    for row in summary_rows:
        row["Section"] = section_str.get(row[tan_key], "")

    summary_df = pd.DataFrame(summary_rows)
    transaction_df = pd.DataFrame(transaction_rows)
    return summary_df, transaction_df


def amount_column_for(df: pd.DataFrame) -> str:
    """Pick the primary 'amount' column to total for the Annual Tax
    Statement Summary sheet, based on naming convention rather than a
    hardcoded per-Part list."""
    if df is None or df.empty:
        return None
    preferred_keywords = ["Amount Paid", "Refund Amount", "Total Transaction Amount", "Total Amount"]
    for kw in preferred_keywords:
        for c in df.columns:
            if kw.lower() in c.lower() and pd.api.types.is_numeric_dtype(df[c]):
                return c
    for c in df.columns:
        if pd.api.types.is_numeric_dtype(df[c]):
            return c
    return None


# ==============================================================
# MAIN APPLICATION
# ==============================================================

if uploaded_file:

    with st.spinner("Reading PDF…"):
        with pdfplumber.open(uploaded_file) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

    # ------------------------------------------------------------------
    # Assessee Details
    # ------------------------------------------------------------------
    assessee = {}
    pan_match = re.search(r"Permanent Account Number \(PAN\)\s+([A-Z0-9]+)", full_text)
    name_match = re.search(r"Name of Assessee\s+(.+?)\n", full_text)
    fy_match = re.search(r"Financial Year\s+([0-9\-]+)", full_text)
    ay_match = re.search(r"Assessment Year\s+([0-9\-]+)", full_text)
    status_match = re.search(r"Current Status of PAN\s+(.+?)\s+Financial Year", full_text)
    address_match = re.search(r"Address of Assessee\s+(.+?)Above data", full_text, re.S)

    assessee["PAN"] = pan_match.group(1) if pan_match else ""
    assessee["Name"] = name_match.group(1).strip() if name_match else ""
    assessee["Financial Year"] = fy_match.group(1) if fy_match else ""
    assessee["Assessment Year"] = ay_match.group(1) if ay_match else ""
    assessee["PAN Status"] = status_match.group(1).strip() if status_match else ""
    assessee["Address"] = address_match.group(1).replace("\n", " ").strip() if address_match else ""

    assessee_df = pd.DataFrame(list(assessee.items()), columns=["Field", "Value"])

    # ------------------------------------------------------------------
    # Dynamic Part detection
    # ------------------------------------------------------------------
    clean_text = strip_noise(full_text, assessee.get("Name", ""))
    parts = find_parts(clean_text)
    contact_lookup = parse_contact_info(clean_text)

    part_results = []  # one entry per detected Part, in document order

    for p in parts:
        roman = p["roman"]
        no_trx = has_no_transactions(p["text"])

        if is_deductor_collector_shape(p["text"]):
            summary_df, transaction_df = parse_deductor_collector_shape(p["text"])
        else:
            summary_df, transaction_df = pd.DataFrame(), pd.DataFrame()

        part_results.append({
            "roman": roman,
            "label": p["label"],
            "description": p["description"],
            "contact": contact_lookup.get(roman, ""),
            "summary_df": summary_df,
            "transaction_df": transaction_df,
            "no_transactions": no_trx and transaction_df.empty,
        })

    # ==================================================================
    # STREAMLIT PREVIEW
    # ==================================================================

    def show_section(title: str, df: pd.DataFrame, no_trx: bool):
        st.subheader(title)
        if df is not None and not df.empty:
            st.dataframe(df, use_container_width=True)
        elif no_trx:
            st.info("No Transactions Present")

    st.subheader("Assessee Details")
    st.dataframe(assessee_df, use_container_width=True)

    for pr in part_results:
        st.markdown(f"### {pr['label']} — {pr['description']}")
        show_section(f"{pr['label']} Summary", pr["summary_df"], pr["no_transactions"])
        show_section(f"{pr['label']} Transactions", pr["transaction_df"], pr["no_transactions"])

    # ==================================================================
    # EXCEL EXPORT
    # ==================================================================

    output = BytesIO()
    _table_counter[0] = 0

    # ---- Build "Annual Tax Statement Summary" sheet rows -------------
    overview_rows = []
    for pr in part_results:
        n_records = len(pr["summary_df"]) if pr["summary_df"] is not None else 0
        amt_col = amount_column_for(pr["summary_df"])
        total_amount = safe_float(pr["summary_df"][amt_col].sum()) if amt_col else 0.0
        overview_rows.append([
            pr["label"],
            pr["description"],
            pr["contact"],
            n_records,
            total_amount,
        ])
    overview_df = pd.DataFrame(overview_rows, columns=[
        "Part", "Description", "Contact", "Number of Records", "Total Amount",
    ])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ---- Sheet 1: Annual Tax Statement Summary --------------------
        overview_df.to_excel(writer, sheet_name="Annual Tax Statement Summary", index=False)
        ws = writer.sheets["Annual Tax Statement Summary"]
        total_pairs = [(5, get_column_letter(5))]
        apply_sheet_formatting(
            ws,
            numeric_cols=[5],
            total_col_indices=total_pairs,
            table_name=_next_table_name("AnnualSummary"),
        )

        # ---- Per-Part Summary + Transactions sheets --------------------
        all_trx_frames = []

        for pr in part_results:
            summary_sheet_name = safe_sheet_name(f"{pr['label']} Summary")
            trx_sheet_name = safe_sheet_name(f"{pr['label']} Transactions")

            sdf = pr["summary_df"]
            num_cols_s = [c for c in (sdf.columns if sdf is not None and not sdf.empty else [])
                          if pd.api.types.is_numeric_dtype(sdf[c]) and c != "S.No"]
            write_df_sheet(
                writer, sdf, summary_sheet_name,
                numeric_col_names=num_cols_s,
                add_totals=bool(num_cols_s),
                table_prefix=safe_sheet_name(f"{pr['label']}Sum"),
                placeholder_message="No Transactions Present",
            )

            tdf = pr["transaction_df"]
            num_cols_t = [c for c in (tdf.columns if tdf is not None and not tdf.empty else [])
                          if pd.api.types.is_numeric_dtype(tdf[c]) and c != "S.No"]
            write_df_sheet(
                writer, tdf, trx_sheet_name,
                numeric_col_names=num_cols_t,
                add_totals=False,
                table_prefix=safe_sheet_name(f"{pr['label']}Trx"),
                placeholder_message="No Transactions Present",
            )

            if tdf is not None and not tdf.empty:
                labeled = tdf.copy()
                labeled.insert(0, "Part", pr["label"])
                all_trx_frames.append(labeled)

        # ---- Final "All Transactions" master register -------------------
        if all_trx_frames:
            all_trx_df = pd.concat(all_trx_frames, ignore_index=True, sort=False)
        else:
            all_trx_df = pd.DataFrame()

        num_cols_all = [c for c in (all_trx_df.columns if not all_trx_df.empty else [])
                        if pd.api.types.is_numeric_dtype(all_trx_df[c]) and c != "S.No"]
        write_df_sheet(
            writer, all_trx_df, "All Transactions",
            numeric_col_names=num_cols_all,
            add_totals=False,
            table_prefix="AllTransactions",
            placeholder_message="No Transactions Present",
        )

    output.seek(0)

    pan = assessee.get("PAN", "").strip()
    name = re.sub(r'[\\/:*?"<>|]', "", assessee.get("Name", "").strip())
    ay = assessee.get("Assessment Year", "").strip()
    output_filename = f"{pan}_{name}_{ay}.xlsx"

    st.download_button(
        "⬇️  Download Excel",
        output,
        file_name=output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
